import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def create_modified_workbook(input_filename, output_filename):
    # Load the existing workbook
    wb = xl.load_workbook(input_filename)
    sheet = wb['Sheet1']

    # Create a new workbook
    new_wb = xl.Workbook()
    new_sheet = new_wb.active

    # Copy the header row to the new workbook
    header_row = sheet[1]
    for col_num, cell in enumerate(header_row, 1):
        new_sheet.cell(row=1, column=col_num, value=cell.value)

    for row in range(2, sheet.max_row + 1):
        transaction_id = sheet.cell(row, 1).value
        product_id = sheet.cell(row, 2).value
        price_with_symbol = sheet.cell(row, 3).value

        # Check if the price is a string with a dollar sign
        if isinstance(price_with_symbol, str) and '$' in price_with_symbol:
            # Remove the dollar sign and convert to float
            price = float(price_with_symbol.replace('$', ''))
        else:
            # Assume the price is already a numeric value
            price = price_with_symbol

        correct_price = price * 0.9

        # Format the prices with a dollar sign
        price_with_symbol = f"${price:,.2f}"  # Retain currency formatting
        correct_price_with_symbol = f"${correct_price:,.2f}"

        # Add values to the new workbook
        new_sheet.append([transaction_id, product_id, price_with_symbol, correct_price])

    values = Reference(new_sheet,
                       min_col=4,
                       min_row=1,
                       max_col=4,
                       max_row=new_sheet.max_row)

    chart = BarChart()
    chart.add_data(values, titles_from_data=True)
    chart.title = "Corrected Prices Chart"
    chart.x_axis.title = "Transaction ID"
    chart.y_axis.title = "Price"
    chart.style = 13
    chart.x_axis.majorTickMark = 'out'
    chart.x_axis.labelRotation = 45

    # Adjust the chart category axis to use transaction IDs from the first column
    categories = Reference(new_sheet,
                           min_col=1,
                           min_row=2,
                           max_row=new_sheet.max_row)
    chart.set_categories(categories)

    # Format the Y-axis labels with dollar sign
    chart.y_axis.majorTickMark = 'out'
    chart.y_axis.labelNumberFormat = '[$$-en-US]#,##0.00'

    # Add the chart to the worksheet
    new_sheet.add_chart(chart, "E2")

    # Save the new workbook with the provided output filename
    new_wb.save(output_filename)

# Ask the user for the existing filename
input_filename = input("Enter the existing filename (including extension): ")
# Ask the user for the new filename
output_filename = input("Enter the new filename (including extension) for the modified workbook: ")

# Call the function to create a new workbook with modifications
create_modified_workbook(input_filename, output_filename)
